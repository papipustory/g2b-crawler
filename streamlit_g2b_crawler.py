import streamlit as st
import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment
import logging
import sys
import subprocess

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Streamlit 페이지 설정
st.set_page_config(
    page_title="💻 나라장터 제안공고 크롤러",
    page_icon="💻",
    layout="centered"
)

st.title("💻 나라장터 제안공고 크롤러")
st.markdown("컴퓨터 관련 제안공고를 G2B에서 크롤링하여 다운로드합니다.")

# 초기화 함수
@st.cache_resource
def initialize_playwright():
    """Playwright 초기 설정"""
    try:
        # Streamlit Cloud 환경 확인
        if os.environ.get('STREAMLIT_SHARING_MODE'):
            # 브라우저 경로 설정
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = '0'
            
            # Chromium 설치 확인
            result = subprocess.run(
                [sys.executable, "-m", "playwright", "install", "chromium"],
                capture_output=True,
                text=True
            )
            logger.info(f"Chromium install output: {result.stdout}")
            
            # 의존성 설치
            result = subprocess.run(
                [sys.executable, "-m", "playwright", "install-deps", "chromium"],
                capture_output=True,
                text=True
            )
            logger.info(f"Dependencies install output: {result.stdout}")
            
        return True
    except Exception as e:
        logger.error(f"Playwright 초기화 실패: {str(e)}")
        return False

# Playwright 초기화
if 'playwright_ready' not in st.session_state:
    with st.spinner("브라우저 환경 준비 중..."):
        st.session_state.playwright_ready = initialize_playwright()

if not st.session_state.playwright_ready:
    st.error("브라우저 초기화에 실패했습니다. 페이지를 새로고침해주세요.")
    st.stop()

# 진행 상태 표시
progress_bar = st.progress(0)
status_text = st.empty()

def update_progress(progress, message):
    """진행 상태 업데이트"""
    progress_bar.progress(progress)
    status_text.text(message)

# PC 버전 크롤링 함수들을 그대로 가져옴
async def close_notice_popups(page):
    """팝업창 닫기 - PC 버전과 동일"""
    for _ in range(5):
        popup_divs = await page.query_selector_all("div[id^='mf_wfm_container_wq_uuid_'][class*='w2popup_window']")
        closed = False
        for popup in popup_divs:
            try:
                # 닫기 버튼 찾기
                for sel in ["button[class*='w2window_close']", "input[type='button'][value='닫기']"]:
                    btn = await popup.query_selector(sel)
                    if btn:
                        await btn.click()
                        await asyncio.sleep(0.2)
                        closed = True
                        break
                
                # 오늘 하루 체크박스 처리
                checkbox = await popup.query_selector("input[type='checkbox'][title*='오늘 하루']")
                if checkbox:
                    await checkbox.check()
                    await asyncio.sleep(0.1)
                    btn = await popup.query_selector("input[type='button'][value='닫기']")
                    if btn:
                        await btn.click()
                        closed = True
                        break
            except:
                continue
        
        if not closed:
            break
        await asyncio.sleep(0.5)

async def wait_and_click(page, selector, desc, timeout=3000, scroll=True):
    """요소 대기 및 클릭 - PC 버전과 동일"""
    try:
        await page.wait_for_selector(selector, timeout=timeout, state="visible")
        elem = await page.query_selector(selector)
        if elem and await elem.is_visible():
            if scroll:
                await elem.scroll_into_view_if_needed()
                await asyncio.sleep(0.02)
            await elem.click()
            return True
        return False
    except:
        return False

async def run_crawler():
    """메인 크롤링 함수"""
    browser = None
    
    try:
        update_progress(0.1, "브라우저 시작 중...")
        
        async with async_playwright() as p:
            # 브라우저 실행 옵션 - Streamlit Cloud에 맞게 수정
            launch_args = [
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-dev-shm-usage',
                '--disable-accelerated-2d-canvas',
                '--no-first-run',
                '--no-zygote',
                '--single-process',
                '--disable-gpu',
                '--disable-blink-features=AutomationControlled'
            ]
            
            browser = await p.chromium.launch(
                headless=True,
                args=launch_args
            )
            
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            )
            
            page = await context.new_page()
            
            update_progress(0.2, "나라장터 접속 중...")
            
            # 페이지 로드
            await page.goto("https://shop.g2b.go.kr/", timeout=30000)
            await page.wait_for_load_state('networkidle', timeout=5000)
            await asyncio.sleep(3)
            
            update_progress(0.3, "팝업 처리 중...")
            
            # 팝업 닫기
            await close_notice_popups(page)
            
            # 페이지 하단 스크롤
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(1)
            
            update_progress(0.4, "제안공고목록 찾는 중...")
            
            # 제안공고목록 버튼 클릭 - PC 버전과 동일한 로직
            btn_selectors = [
                'a[id^="mf_wfm_container_wq_uuid_"][id$="_btnPrpblist"]',
                'a[title*="제안공고목록"]',
                'a:has-text("제안공고목록")',
                '//a[contains(text(), "제안공고목록")]',
                'div.w2textbox:text("제안공고목록")',
            ]
            
            clicked = False
            for sel in btn_selectors:
                if await wait_and_click(page, sel, "제안공고목록 버튼"):
                    clicked = True
                    break
            
            if not clicked:
                # 모든 a 태그 검사
                all_a = await page.query_selector_all("a")
                for a in all_a:
                    try:
                        title = await a.get_attribute("title")
                        href = await a.get_attribute("href")
                        inner = await a.inner_text()
                        if (title and "제안공고" in title) or (inner and "제안공고" in inner):
                            if href and href.strip() != "javascript:void(null)":
                                await a.scroll_into_view_if_needed()
                                await asyncio.sleep(0.2)
                                await a.click()
                                clicked = True
                                break
                    except:
                        continue
            
            await asyncio.sleep(3)
            
            update_progress(0.5, "검색 조건 설정 중...")
            
            # 3개월 라디오 버튼 선택
            await page.evaluate("""
                const radio = document.querySelector('input[title="3개월"]');
                if (radio) {
                    radio.checked = true;
                    const event = new Event('click', { bubbles: true });
                    radio.dispatchEvent(event);
                }
            """)
            await asyncio.sleep(1)
            
            update_progress(0.6, "검색어 입력 중...")
            
            # 검색어 입력
            input_elem = await page.query_selector('td[data-title="제안공고명"] input[type="text"]')
            if input_elem:
                await input_elem.fill('컴퓨터', timeout=10000)
            
            # 표시수 100건으로 변경
            await page.evaluate("""
                const selects = document.querySelectorAll('select[id*="RecordCountPerPage"]');
                selects.forEach(select => {
                    select.value = "100";
                    const event = new Event('change', { bubbles: true });
                    select.dispatchEvent(event);
                });
            """)
            await asyncio.sleep(1)
            
            update_progress(0.7, "검색 실행 중...")
            
            # 적용 및 검색 버튼 클릭
            await wait_and_click(page, 'input[type="button"][value="적용"]', "적용버튼", scroll=False)
            await asyncio.sleep(0.5)
            await wait_and_click(page, 'input[type="button"][value="검색"]', "검색버튼", scroll=False)
            await asyncio.sleep(3)
            
            update_progress(0.8, "데이터 수집 중...")
            
            # 테이블 데이터 수집
            table_elem = await page.query_selector('table[id$="grdPrpsPbanc_body_table"]')
            data = []
            
            if table_elem:
                rows = await table_elem.query_selector_all('tr')
                
                for row in rows:
                    tds = await row.query_selector_all('td')
                    cols = []
                    
                    for td in tds:
                        # nobr 태그 확인
                        nobr = await td.query_selector('nobr')
                        if nobr:
                            text = await nobr.inner_text()
                        else:
                            # a 태그 확인
                            a = await td.query_selector('a')
                            if a:
                                text = await a.inner_text()
                            else:
                                text = await td.inner_text()
                        
                        cols.append(text.strip())
                    
                    if cols and any(cols):
                        data.append(cols)
            
            update_progress(0.9, "데이터 저장 중...")
            
            # 데이터 처리 및 저장
            if data:
                new_df = pd.DataFrame(data)
                
                # 헤더 설정
                headers = ["No", "제안공고번호", "수요기관", "제안공고명", "공고게시일자", "공고마감일시", "공고상태", "사유", "기타"]
                if len(new_df.columns) < len(headers):
                    headers = headers[:len(new_df.columns)]
                elif len(new_df.columns) > len(headers):
                    new_df = new_df.iloc[:, :len(headers)]
                
                new_df.columns = headers
                
                # 기존 파일과 병합
                file_path = 'g2b_result.xlsx'
                if os.path.exists(file_path):
                    old_df = pd.read_excel(file_path)
                    combined_df = pd.concat([old_df, new_df])
                    combined_df.drop_duplicates(subset="제안공고번호", keep='last', inplace=True)
                    combined_df.reset_index(drop=True, inplace=True)
                else:
                    combined_df = new_df
                
                # Excel 저장
                combined_df.to_excel(file_path, index=False)
                
                # Excel 서식 적용
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # 정렬 설정
                align = Alignment(horizontal='center', vertical='center')
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = align
                
                # 열 너비 설정
                col_widths = [3.5, 17, 44, 55, 15, 17.5, 17, 17, 17]
                for i, width in enumerate(col_widths, start=1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
                
                wb.save(file_path)
                
                update_progress(1.0, "크롤링 완료!")
                return len(data)
            else:
                update_progress(1.0, "검색 결과가 없습니다.")
                return 0
            
    except Exception as e:
        logger.error(f"크롤링 오류: {str(e)}")
        st.error(f"오류 발생: {str(e)}")
        return -1
    finally:
        if browser:
            await browser.close()

# 메인 UI
col1, col2 = st.columns([3, 1])

with col1:
    search_button = st.button("🚀 크롤링 시작", type="primary", use_container_width=True)

with col2:
    if os.path.exists('g2b_result.xlsx'):
        with open('g2b_result.xlsx', 'rb') as f:
            st.download_button(
                label="📥 기존 파일",
                data=f,
                file_name="g2b_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# 크롤링 실행
if search_button:
    with st.spinner("크롤링 진행 중... (1-2분 소요)"):
        # asyncio 실행
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(run_crawler())
        
        if result > 0:
            st.success(f"✅ 크롤링 완료! 총 {result}개의 공고를 수집했습니다.")
            
            # 다운로드 버튼
            with open('g2b_result.xlsx', 'rb') as f:
                st.download_button(
                    label="📥 결과 다운로드 (Excel)",
                    data=f,
                    file_name="g2b_result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            # 데이터 미리보기
            df = pd.read_excel('g2b_result.xlsx')
            st.markdown("### 📊 수집된 데이터 미리보기")
            st.dataframe(df.head(10), use_container_width=True)
            
        elif result == 0:
            st.warning("검색 결과가 없습니다.")
        else:
            st.error("크롤링 중 오류가 발생했습니다.")

# 추가 정보
with st.expander("ℹ️ 사용 안내"):
    st.markdown("""
    ### 크롤링 정보
    - **검색어**: 컴퓨터
    - **검색 기간**: 3개월
    - **표시 건수**: 100건
    - **중복 제거**: 제안공고번호 기준
    
    ### 주의사항
    - 나라장터 시스템 점검 시간에는 작동하지 않습니다
    - 네트워크 상태에 따라 시간이 걸릴 수 있습니다
    - 기존 파일이 있으면 자동으로 병합됩니다
    """)

# 사이드바
with st.sidebar:
    st.markdown("### 🔧 크롤러 설정")
    st.info("현재 검색어: 컴퓨터")
    st.info("검색 기간: 3개월")
    st.info("표시 건수: 100건")
    
    st.markdown("---")
    
    if st.button("🗑️ 기존 데이터 삭제"):
        if os.path.exists('g2b_result.xlsx'):
            os.remove('g2b_result.xlsx')
            st.success("기존 데이터가 삭제되었습니다.")
            st.rerun()
